"""XLSX import parsing: forward-filled dates, "-" handling, idempotency."""

from __future__ import annotations

import io

from openpyxl import Workbook

from app.services import xlsx_import as service
from app.services.xlsx_import import parse_duration_minutes, parse_excel_date


def _build_workbook() -> bytes:
    wb = Workbook()
    daily = wb.active
    daily.title = "Daily Tracker"
    daily.append(["Piriformis Recovery Tracker"])  # row 1 — title
    daily.append([])  # row 2 — blank
    daily.append(
        [
            "Date",
            "Day",
            "Status",
            "Strengthening Session?",
            "Session Intensity",
            "Sharp Pain Episodes",
            "Worst Pain",
            "Tingling Level",
            "Tingling Duration",
            "Stretches Done?",
            "Sitting Breaks Taken?",
            "Sleep Quality",
            "Notes",
        ]
    )  # row 3 — header
    daily.append(
        ["2026-06-12", "Fri", "G", "Yes", "6", "2", "3", "4", "4hrs", "Yes", "Yes - Many", "4", "Felt ok"]
    )
    daily.append(
        ["2026-06-13", "Sat", "A", "No", "-", "0", "-", "2", "30min", "No", "A few", "3", "Rest day"]
    )

    log = wb.create_sheet("Exercise Log")
    log.append(["Strength Log"])
    log.append([])
    log.append(
        ["Date", "Exercise", "Sets", "Reps", "Modification/Progression", "Difficulty", "Nerve Response During?", "Notes"]
    )
    # Session 1 on 2026-06-12: date only on first row (forward-fill).
    log.append(["2026-06-12", "Glute Bridges", "3", "12", "", "4", "none", ""])
    log.append(["", "Clamshells", "3", "15", "band", "5", "slight twinge", ""])
    log.append(["", "Forearm Plank", "1", "45", "", "6", "", ""])  # seconds -> hold_seconds
    log.append(["", "Step-Ups", "-", "-", "", "", "skipped", ""])  # not performed -> skip
    # Session 2 on 2026-06-14.
    log.append(["2026-06-14", "Goblet Squats", "3", "10", "8kg", "5", "", ""])

    weekly = wb.create_sheet("Weekly Summary")
    weekly.append(
        ["Week", "Date Range", "Strengthening Sessions", "Avg Pain Episodes/Day", "Avg Tingling Level", "Worst Pain Day", "Overall Status", "Exercises Completed", "Pigeon Pose Progress?", "Key Observations", "Trend vs Last Week"]
    )
    weekly.append(
        ["Week 1", "12/06/2026 - 18/06/2026", "2", "1.0", "3.0", "5", "A", "all", "yes", "Good progress overall", "Better"]
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_duration_minutes():
    assert parse_duration_minutes("4hrs") == (240, None)
    assert parse_duration_minutes("30min") == (30, None)
    assert parse_duration_minutes("1.5hr") == (90, None)
    assert parse_duration_minutes("-") == (None, None)
    mins, raw = parse_duration_minutes("all day")
    assert mins is None and raw == "all day"


def test_parse_excel_date_serial():
    from datetime import date

    # Round-trip a known date through its 1900-system serial number.
    serial = (date(2026, 6, 12) - date(1899, 12, 30)).days
    assert parse_excel_date(serial) == date(2026, 6, 12)


def test_import_full_workbook(db, user_id):
    content = _build_workbook()
    result = service.import_workbook(db, user_id, content)
    assert result["daily_entries"] == 2
    assert result["sessions"] == 2
    assert result["weekly_summaries"] == 1

    # Daily values parsed, "-" -> NULL, duration parsed, stretches split.
    row = db.query_one("SELECT * FROM daily_entries WHERE entry_date = '2026-06-12'")
    assert row["status"] == "G"
    assert row["tingling_duration_minutes"] == 240
    assert row["stretches_morning"] and row["stretches_night"]
    rest = db.query_one("SELECT * FROM daily_entries WHERE entry_date = '2026-06-13'")
    assert rest["session_intensity"] is None
    assert rest["worst_pain"] is None
    assert rest["tingling_duration_minutes"] == 30

    # Forward-filled session: 3 performed exercises (Step-Ups skipped).
    logs = db.query(
        """
        SELECT e.name, el.reps, el.hold_seconds
        FROM exercise_logs el
        JOIN strength_sessions s ON s.id = el.session_id
        JOIN daily_entries d ON d.id = s.daily_entry_id
        JOIN exercises e ON e.id = el.exercise_id
        WHERE d.entry_date = '2026-06-12'
        """
    )
    names = {r["name"] for r in logs}
    assert names == {"Glute Bridges", "Clamshells", "Forearm Plank"}
    plank = next(r for r in logs if r["name"] == "Forearm Plank")
    assert plank["hold_seconds"] == 45 and plank["reps"] is None


def test_import_is_idempotent(db, user_id):
    content = _build_workbook()
    service.import_workbook(db, user_id, content)
    service.import_workbook(db, user_id, content)
    # 06-12 + 06-13 from the daily sheet, plus 06-14 created by the exercise log.
    assert db.query_one("SELECT COUNT(*) AS n FROM daily_entries")["n"] == 3
    assert db.query_one("SELECT COUNT(*) AS n FROM strength_sessions")["n"] == 2
    # No duplicated exercise logs for the forward-filled session.
    n_logs = db.query_one(
        "SELECT COUNT(*) AS n FROM exercise_logs el "
        "JOIN strength_sessions s ON s.id = el.session_id "
        "JOIN daily_entries d ON d.id = s.daily_entry_id "
        "WHERE d.entry_date = '2026-06-12'"
    )["n"]
    assert n_logs == 3


def test_import_writes_notes(db, user_id):
    """Notes column is stored in the notes table; re-import is idempotent."""
    content = _build_workbook()

    # First import: a notes row should be created for each row that has note text.
    service.import_workbook(db, user_id, content)
    note_row = db.query_one(
        "SELECT n.body, n.source "
        "FROM notes n "
        "JOIN daily_entries d ON d.id = n.daily_entry_id "
        "WHERE d.entry_date = '2026-06-12' AND n.source = 'import'"
    )
    assert note_row is not None
    assert note_row["source"] == "import"
    assert "Felt ok" in note_row["body"]

    # Re-import: still exactly ONE import-sourced note for that day.
    service.import_workbook(db, user_id, content)
    count = db.query_one(
        "SELECT COUNT(*) AS n "
        "FROM notes n "
        "JOIN daily_entries d ON d.id = n.daily_entry_id "
        "WHERE d.entry_date = '2026-06-12' AND n.source = 'import'"
    )["n"]
    assert count == 1
