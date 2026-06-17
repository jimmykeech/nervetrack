"""Spreadsheet import.

Imports the three sheets of the legacy "Piriformis Recovery Tracker" workbook.
Tolerant of messy values ("-", blanks, "Yes - Many", trailing spaces). Idempotent:
re-running matches on date (daily/weekly) and date+exercise (logs) without
duplicating.
"""

from __future__ import annotations

import re
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any
from uuid import UUID

from openpyxl import load_workbook

from app.db import Database
from app.services.timeutil import now_utc

# Exercises whose "Reps" cell actually records a hold time in seconds.
_TIME_BASED = {"forearm plank", "hollowbody hold", "side plank", "hollow body hold"}


# --- value coercion helpers ------------------------------------------------


def _clean(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _is_blank(v: Any) -> bool:
    return _clean(v) in ("", "-", "–", "—", "n/a", "na")


def parse_excel_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial date (1900 system): day 1 == 1899-12-31, with the
        # well-known 1900 leap-year bug offset.
        return (datetime(1899, 12, 30) + timedelta(days=float(v))).date()
    s = _clean(v)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_decimal(v: Any) -> Decimal | None:
    if _is_blank(v):
        return None
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return None


def parse_int(v: Any) -> int | None:
    d = parse_decimal(v)
    return int(d) if d is not None else None


def parse_yesno(v: Any) -> bool:
    return _clean(v).lower().startswith("y")


def parse_duration_minutes(v: Any) -> tuple[int | None, str | None]:
    """Parse free-text durations like "4hrs", "30min", "1.5hr" to minutes.

    Returns (minutes, leftover_raw). leftover_raw is the original string when it
    could not be parsed, so callers can preserve it in notes.
    """
    if _is_blank(v):
        return None, None
    s = _clean(v).lower()
    total = 0.0
    matched = False
    pattern = r"(\d+(?:\.\d+)?)\s*(h|hr|hrs|hour|hours|m|min|mins|minutes)?"
    for value, unit in re.findall(pattern, s):
        if not value:
            continue
        num = float(value)
        if unit in ("", "h", "hr", "hrs", "hour", "hours") and ("m" not in unit):
            # bare number or hour units -> treat as hours
            total += num * 60
        else:
            total += num
        matched = True
    if not matched:
        return None, s
    return int(round(total)), None


# --- sheet locators --------------------------------------------------------


def _find_sheet(wb, *needles: str):
    for ws in wb.worksheets:
        title = ws.title.strip().lower()
        if any(n in title for n in needles):
            return ws
    return None


def _header_index(ws, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col, cell in enumerate(ws[header_row]):
        name = _clean(cell.value).lower()
        if name:
            headers[name] = col
    return headers


def _get(row: tuple, headers: dict[str, int], *needles: str) -> Any:
    for key, idx in headers.items():
        if any(n in key for n in needles):
            if idx < len(row):
                return row[idx]
    return None


# --- importers -------------------------------------------------------------


def _import_daily(db: Database, user_id: UUID, ws) -> int:
    headers = _header_index(ws, 3)
    if not headers:
        return 0
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        d = parse_excel_date(_get(row, headers, "date"))
        if d is None:
            continue
        dur_min, dur_raw = parse_duration_minutes(
            _get(row, headers, "tingling duration", "duration")
        )
        notes = _clean(_get(row, headers, "notes")) or None
        if dur_raw:
            tag = f"[tingling duration: {dur_raw}]"
            notes = f"{notes}\n{tag}" if notes else tag
        stretches = parse_yesno(_get(row, headers, "stretches"))
        status = _clean(_get(row, headers, "status")).upper()[:1]
        status = status if status in ("G", "A", "R") else None
        with db.cursor():
            db.execute(
                """
                INSERT INTO daily_entries
                    (user_id, entry_date, status, strengthening_done, session_intensity,
                     sharp_pain_episodes, worst_pain, tingling_level,
                     tingling_duration_minutes, stretches_morning, stretches_night,
                     sitting_breaks, sleep_quality, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (user_id, entry_date) DO UPDATE SET
                    status = excluded.status,
                    strengthening_done = excluded.strengthening_done,
                    session_intensity = excluded.session_intensity,
                    sharp_pain_episodes = excluded.sharp_pain_episodes,
                    worst_pain = excluded.worst_pain,
                    tingling_level = excluded.tingling_level,
                    tingling_duration_minutes = excluded.tingling_duration_minutes,
                    stretches_morning = excluded.stretches_morning,
                    stretches_night = excluded.stretches_night,
                    sitting_breaks = excluded.sitting_breaks,
                    sleep_quality = excluded.sleep_quality,
                    updated_at = excluded.updated_at
                """,
                [
                    user_id,
                    d,
                    status,
                    parse_yesno(_get(row, headers, "strengthening")),
                    parse_decimal(_get(row, headers, "intensity")),
                    parse_int(_get(row, headers, "sharp pain", "episodes")) or 0,
                    parse_decimal(_get(row, headers, "worst pain")),
                    parse_decimal(_get(row, headers, "tingling level")),
                    dur_min,
                    stretches,
                    stretches,
                    _clean(_get(row, headers, "sitting breaks", "breaks")) or None,
                    parse_decimal(_get(row, headers, "sleep")),
                    now_utc(),
                ],
            )
        if notes:
            entry_row = db.query_one(
                "SELECT id FROM daily_entries WHERE user_id = ? AND entry_date = ?",
                [user_id, d],
            )
            with db.cursor():
                db.execute(
                    "DELETE FROM notes WHERE daily_entry_id = ? AND source = 'import'",
                    [entry_row["id"]],
                )
                db.execute(
                    "INSERT INTO notes (daily_entry_id, occurred_at, body, source) "
                    "VALUES (?, ?, ?, 'import')",
                    [entry_row["id"], datetime(d.year, d.month, d.day, 12, 0), notes],
                )
        count += 1
    return count


def _parse_range_start(v: Any) -> date | None:
    s = _clean(v)
    if not s:
        return None
    first = re.split(r"[-–—]", s)[0].strip()
    return parse_excel_date(first)


def _import_weekly(db: Database, user_id: UUID, ws) -> int:
    # Header is on the first non-empty row; scan the first few rows for it.
    header_row = 1
    for r in range(1, 5):
        idx = _header_index(ws, r)
        if any("week" in k for k in idx) or any("date range" in k for k in idx):
            header_row = r
            break
    headers = _header_index(ws, header_row)
    if not headers:
        return 0
    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        wk = _parse_range_start(_get(row, headers, "date range", "range"))
        if wk is None:
            continue
        status = _clean(_get(row, headers, "overall status", "status")).upper()[:1]
        status = status if status in ("G", "A", "R") else None
        trend = _clean(_get(row, headers, "trend")) or None
        observations = _clean(_get(row, headers, "key observations", "observations")) or None
        with db.cursor():
            db.execute(
                """
                INSERT INTO weekly_summaries
                    (user_id, week_start, overall_status, key_observations, trend_vs_last_week)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT (user_id, week_start) DO UPDATE SET
                    overall_status = excluded.overall_status,
                    key_observations = excluded.key_observations,
                    trend_vs_last_week = excluded.trend_vs_last_week
                """,
                [user_id, wk, status, observations, trend],
            )
        count += 1
    return count


def _resolve_exercise(db: Database, user_id: UUID, name: str) -> Any:
    clean = name.strip()
    row = db.query_one(
        "SELECT id FROM exercises WHERE user_id = ? AND lower(name) = lower(?)",
        [user_id, clean],
    )
    if row:
        return row["id"]
    created = db.query_one(
        "INSERT INTO exercises (user_id, name, active, sort_order) "
        "VALUES (?, ?, TRUE, 999) RETURNING id",
        [user_id, clean],
    )
    return created["id"]


def _import_exercise_log(db: Database, user_id: UUID, ws) -> int:
    headers = _header_index(ws, 3)
    if not headers:
        return 0
    # Group rows into sessions by forward-filled date.
    sessions: dict[date, list[dict]] = {}
    current_date: date | None = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        raw_date = _get(row, headers, "date")
        if not _is_blank(raw_date):
            parsed = parse_excel_date(raw_date)
            if parsed is not None:
                current_date = parsed
        if current_date is None:
            continue
        name = _clean(_get(row, headers, "exercise"))
        if not name:
            continue
        sets = parse_int(_get(row, headers, "sets"))
        reps_raw = _get(row, headers, "reps")
        # Rows with no sets/reps mean the exercise was not performed — skip.
        if (sets is None or sets == 0) and _is_blank(reps_raw):
            continue
        log: dict[str, Any] = {
            "name": name,
            "sets": sets,
            "difficulty": parse_decimal(_get(row, headers, "difficulty")),
            "nerve_response": _clean(_get(row, headers, "nerve response", "nerve")) or None,
            "modification": _clean(_get(row, headers, "modification", "progression")) or None,
        }
        reps_val = parse_int(reps_raw)
        if name.lower() in _TIME_BASED or "plank" in name.lower() or "hold" in name.lower():
            log["hold_seconds"] = reps_val
            log["reps"] = None
        else:
            log["reps"] = reps_val
            log["hold_seconds"] = None
        sessions.setdefault(current_date, []).append(log)

    count = 0
    for sess_date, logs in sessions.items():
        with db.cursor():
            entry = db.query_one(
                "SELECT id FROM daily_entries WHERE user_id = ? AND entry_date = ?",
                [user_id, sess_date],
            )
            entry_id = (
                entry["id"]
                if entry
                else db.query_one(
                    "INSERT INTO daily_entries (user_id, entry_date, strengthening_done) "
                    "VALUES (?, ?, TRUE) RETURNING id",
                    [user_id, sess_date],
                )["id"]
            )
            # Idempotent: replace any existing imported session for this date.
            existing = db.query(
                "SELECT id FROM strength_sessions WHERE daily_entry_id = ?", [entry_id]
            )
            for s in existing:
                db.execute("DELETE FROM exercise_logs WHERE session_id = ?", [s["id"]])
                db.execute("DELETE FROM strength_sessions WHERE id = ?", [s["id"]])
            performed_at = datetime.combine(sess_date, time(12, 0))
            session_id = db.query_one(
                "INSERT INTO strength_sessions (daily_entry_id, performed_at) "
                "VALUES (?, ?) RETURNING id",
                [entry_id, performed_at],
            )["id"]
            for log in logs:
                ex_id = _resolve_exercise(db, user_id, log["name"])
                db.execute(
                    """
                    INSERT INTO exercise_logs
                        (session_id, exercise_id, sets, reps, hold_seconds,
                         difficulty, nerve_response, modification)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    [
                        session_id,
                        ex_id,
                        log["sets"],
                        log.get("reps"),
                        log.get("hold_seconds"),
                        log["difficulty"],
                        log["nerve_response"],
                        log["modification"],
                    ],
                )
            db.execute(
                "UPDATE daily_entries SET strengthening_done = TRUE WHERE id = ?", [entry_id]
            )
        count += 1
    return count


def import_workbook(db: Database, user_id: UUID, content: bytes) -> dict[str, int]:
    import io

    wb = load_workbook(io.BytesIO(content), data_only=True)
    result = {"daily_entries": 0, "weekly_summaries": 0, "sessions": 0}

    daily_ws = _find_sheet(wb, "daily")
    if daily_ws is not None:
        result["daily_entries"] = _import_daily(db, user_id, daily_ws)

    log_ws = _find_sheet(wb, "exercise log", "exercise")
    if log_ws is not None and log_ws is not daily_ws:
        result["sessions"] = _import_exercise_log(db, user_id, log_ws)

    weekly_ws = _find_sheet(wb, "weekly")
    if weekly_ws is not None:
        result["weekly_summaries"] = _import_weekly(db, user_id, weekly_ws)

    return result
